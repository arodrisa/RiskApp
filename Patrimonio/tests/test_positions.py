import os
import tempfile
import unittest
from unittest.mock import patch
from datetime import datetime
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

# Keep tests isolated from the local development .env loaded by app.main.
os.environ['APP_AUTH_ENABLED'] = 'false'
os.environ['APP_RESTORE_ENABLED'] = 'true'

from app import auth, crud, schemas
from app import models
from app.main import app
from app.database import get_db
from fastapi.encoders import jsonable_encoder
from fastapi import HTTPException
from fastapi.testclient import TestClient
from import_excel import detect_name, detect_category, detect_net_asset_value, detect_sheet_date, import_file


def jsonable_backup(payload):
    return jsonable_encoder(payload)


class PositionCrudTests(unittest.TestCase):
    def test_health_endpoint(self):
        client = TestClient(app)
        response = client.get('/health')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {'status': 'ok'})

    def test_auth_guard_requires_login_when_enabled(self):
        previous_enabled = os.environ.get('APP_AUTH_ENABLED')
        previous_user = os.environ.get('PATRIMONIO_USERNAME')
        previous_password = os.environ.get('PATRIMONIO_PASSWORD')
        os.environ['APP_AUTH_ENABLED'] = 'true'
        os.environ['PATRIMONIO_USERNAME'] = 'admin'
        os.environ['PATRIMONIO_PASSWORD'] = 'secret'
        original_user_count = auth._database_user_count
        auth._database_user_count = lambda db=None: 0
        try:
            client = TestClient(app)
            self.assertEqual(client.get('/owners/').status_code, 401)
            login = client.post('/auth/login', json={'username': 'admin', 'password': 'secret'})
            self.assertEqual(login.status_code, 200)
            self.assertTrue(login.json()['authenticated'])
            csrf_token = login.json()['csrf_token']
            self.assertEqual(client.get('/owners/').status_code, 200)
            logout = client.post('/auth/logout', headers={'X-CSRF-Token': csrf_token})
            self.assertEqual(logout.status_code, 200)
            self.assertEqual(client.get('/owners/').status_code, 401)
        finally:
            auth._database_user_count = original_user_count
            auth._LOGIN_ATTEMPTS.clear()
            if previous_enabled is None:
                os.environ.pop('APP_AUTH_ENABLED', None)
            else:
                os.environ['APP_AUTH_ENABLED'] = previous_enabled
            if previous_user is None:
                os.environ.pop('PATRIMONIO_USERNAME', None)
            else:
                os.environ['PATRIMONIO_USERNAME'] = previous_user
            if previous_password is None:
                os.environ.pop('PATRIMONIO_PASSWORD', None)
            else:
                os.environ['PATRIMONIO_PASSWORD'] = previous_password

    def test_csrf_guard_rejects_authenticated_mutation_without_token(self):
        previous_enabled = os.environ.get('APP_AUTH_ENABLED')
        previous_user = os.environ.get('PATRIMONIO_USERNAME')
        previous_password = os.environ.get('PATRIMONIO_PASSWORD')
        os.environ['APP_AUTH_ENABLED'] = 'true'
        os.environ['PATRIMONIO_USERNAME'] = 'admin'
        os.environ['PATRIMONIO_PASSWORD'] = 'secret'
        original_user_count = auth._database_user_count
        auth._database_user_count = lambda db=None: 0
        try:
            client = TestClient(app)
            login = client.post('/auth/login', json={'username': 'admin', 'password': 'secret'})
            self.assertEqual(login.status_code, 200)
            blocked = client.post('/auth/logout')
            self.assertEqual(blocked.status_code, 403)
            allowed = client.post('/auth/logout', headers={'X-CSRF-Token': login.json()['csrf_token']})
            self.assertEqual(allowed.status_code, 200)
        finally:
            auth._database_user_count = original_user_count
            auth._LOGIN_ATTEMPTS.clear()
            if previous_enabled is None:
                os.environ.pop('APP_AUTH_ENABLED', None)
            else:
                os.environ['APP_AUTH_ENABLED'] = previous_enabled
            if previous_user is None:
                os.environ.pop('PATRIMONIO_USERNAME', None)
            else:
                os.environ['PATRIMONIO_USERNAME'] = previous_user
            if previous_password is None:
                os.environ.pop('PATRIMONIO_PASSWORD', None)
            else:
                os.environ['PATRIMONIO_PASSWORD'] = previous_password

    def test_bootstrap_creates_database_user_and_allows_company_ownership(self):
        previous_enabled = os.environ.get('APP_AUTH_ENABLED')
        previous_environment = os.environ.get('APP_ENV')
        previous_token = os.environ.get('APP_BOOTSTRAP_TOKEN')
        os.environ['APP_AUTH_ENABLED'] = 'true'
        os.environ['APP_ENV'] = 'development'
        os.environ.pop('APP_BOOTSTRAP_TOKEN', None)
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)
        db = TestingSession()
        antonio = crud.create_owner(db, 'Antonio', is_family_member=True)
        hereditas = crud.create_owner(db, 'Hereditas', type='company')
        other_company = crud.create_owner(db, 'Other Company', type='company')
        antonio_id = antonio.id
        hereditas_id = hereditas.id
        other_company_id = other_company.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/auth/bootstrap', json={
                'email': 'antonio@example.com',
                'display_name': 'Antonio',
                'password': 'a-long-test-password',
                'person_owner_id': antonio_id,
            })
            self.assertEqual(response.status_code, 200)
            csrf_token = response.json()['csrf_token']
            self.assertEqual(response.json()['role'], 'owner')

            with patch('app.auth._session_context', return_value={'username': 'antonio@example.com', 'role': 'owner'}):
                first_relation = client.post('/entity-ownerships/', headers={'X-CSRF-Token': csrf_token}, json={
                    'owner_id': antonio_id,
                    'owned_id': hereditas_id,
                    'share': 0.5,
                    'effective_from': '2025-01-01',
                })
                self.assertEqual(first_relation.status_code, 200)
                second_relation = client.post('/entity-ownerships/', headers={'X-CSRF-Token': csrf_token}, json={
                    'owner_id': hereditas_id,
                    'owned_id': other_company_id,
                    'share': 0.75,
                    'effective_from': '2025-01-01',
                })
                self.assertEqual(second_relation.status_code, 200)
                cycle = client.post('/entity-ownerships/', headers={'X-CSRF-Token': csrf_token}, json={
                    'owner_id': other_company_id,
                    'owned_id': hereditas_id,
                    'share': 0.2,
                    'effective_from': '2025-01-01',
                })
                self.assertEqual(cycle.status_code, 400)
        finally:
            app.dependency_overrides.clear()
            auth._LOGIN_ATTEMPTS.clear()
            if previous_enabled is None:
                os.environ.pop('APP_AUTH_ENABLED', None)
            else:
                os.environ['APP_AUTH_ENABLED'] = previous_enabled
            if previous_environment is None:
                os.environ.pop('APP_ENV', None)
            else:
                os.environ['APP_ENV'] = previous_environment
            if previous_token is None:
                os.environ.pop('APP_BOOTSTRAP_TOKEN', None)
            else:
                os.environ['APP_BOOTSTRAP_TOKEN'] = previous_token

    def test_password_reset_tokens_are_single_use(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)
        db = TestingSession()
        previous_delivery = os.environ.get('APP_PASSWORD_RESET_DELIVERY')
        previous_public_url = os.environ.get('APP_PUBLIC_URL')
        os.environ['APP_PASSWORD_RESET_DELIVERY'] = 'console'
        os.environ['APP_PUBLIC_URL'] = 'http://127.0.0.1:8001'
        try:
            user = models.User(
                email='antonio@example.com',
                display_name='Antonio',
                password_hash=auth.password_hash('old-password'),
            )
            db.add(user)
            db.commit()

            requested = auth.request_password_reset(
                schemas.PasswordResetRequest(email='antonio@example.com'),
                db,
            )
            self.assertIn('?reset=', requested['dev_reset_url'])
            token = requested['dev_reset_url'].split('?reset=', 1)[1]
            result = auth.confirm_password_reset(
                schemas.PasswordResetConfirm(token=token, password='new-password'),
                db,
            )
            self.assertIn('Password updated', result['message'])
            db.refresh(user)
            self.assertTrue(auth.verify_password('new-password', user.password_hash))
            self.assertEqual(user.session_version, 2)
            with self.assertRaises(Exception):
                auth.confirm_password_reset(
                    schemas.PasswordResetConfirm(token=token, password='another-password'),
                    db,
                )
        finally:
            db.close()
            if previous_delivery is None:
                os.environ.pop('APP_PASSWORD_RESET_DELIVERY', None)
            else:
                os.environ['APP_PASSWORD_RESET_DELIVERY'] = previous_delivery
            if previous_public_url is None:
                os.environ.pop('APP_PUBLIC_URL', None)
            else:
                os.environ['APP_PUBLIC_URL'] = previous_public_url

    def test_viewer_role_cannot_mutate_project_data(self):
        with patch('app.auth._session_context', return_value={'username': 'viewer@example.com', 'role': 'viewer'}):
            with self.assertRaises(HTTPException) as raised:
                auth.require_project_editor(None)
            self.assertEqual(raised.exception.status_code, 403)

        with patch('app.auth._session_context', return_value={'username': 'editor@example.com', 'role': 'editor'}):
            self.assertEqual(auth.require_project_editor(None), 'editor@example.com')

    def test_project_admin_can_change_invited_user_role_and_revoke_access(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)
        db = TestingSession()
        try:
            project = crud.get_default_project(db)
            admin = models.User(email='admin@example.com', display_name='Admin', password_hash=auth.password_hash('admin-password'))
            invited = models.User(email='viewer@example.com', display_name='Viewer', password_hash=auth.password_hash('viewer-password'))
            db.add_all([admin, invited])
            db.flush()
            db.add(models.ProjectMembership(project_id=project.id, user_id=admin.id, role='owner'))
            db.add(models.ProjectMembership(project_id=project.id, user_id=invited.id, role='viewer'))
            db.commit()

            with patch('app.auth._session_context', return_value={
                'username': admin.email,
                'user_id': admin.id,
                'role': 'owner',
                'project_id': project.id,
            }):
                membership = auth.update_project_user(
                    None,
                    invited.id,
                    schemas.ProjectUserUpdate(role='editor', is_active=False),
                    db,
                )
            self.assertEqual(membership.role, 'editor')
            self.assertFalse(membership.user.is_active)
            self.assertEqual(membership.user.session_version, 2)
        finally:
            db.close()

    def test_production_requires_secure_session_configuration(self):
        previous_environment = os.environ.get('APP_ENV')
        previous_secret = os.environ.get('PATRIMONIO_SESSION_SECRET')
        previous_secure = os.environ.get('APP_COOKIE_SECURE')
        previous_public_url = os.environ.get('APP_PUBLIC_URL')
        previous_auth_enabled = os.environ.get('APP_AUTH_ENABLED')
        os.environ['APP_ENV'] = 'production'
        os.environ['APP_AUTH_ENABLED'] = 'true'
        os.environ['PATRIMONIO_SESSION_SECRET'] = 'change-me'
        os.environ['APP_COOKIE_SECURE'] = 'true'
        os.environ['APP_PUBLIC_URL'] = 'https://patrimonio.example.com'
        try:
            with self.assertRaises(RuntimeError):
                auth.validate_production_settings()
            os.environ['PATRIMONIO_SESSION_SECRET'] = 'a' * 40
            auth.validate_production_settings()
        finally:
            for name, value in {
                'APP_ENV': previous_environment,
                'PATRIMONIO_SESSION_SECRET': previous_secret,
                'APP_COOKIE_SECURE': previous_secure,
                'APP_PUBLIC_URL': previous_public_url,
                'APP_AUTH_ENABLED': previous_auth_enabled,
            }.items():
                if value is None:
                    os.environ.pop(name, None)
                else:
                    os.environ[name] = value

    def test_login_rate_limit_blocks_repeated_failures(self):
        previous_enabled = os.environ.get('APP_AUTH_ENABLED')
        previous_user = os.environ.get('PATRIMONIO_USERNAME')
        previous_password = os.environ.get('PATRIMONIO_PASSWORD')
        previous_attempts = os.environ.get('APP_LOGIN_RATE_LIMIT_ATTEMPTS')
        os.environ['APP_AUTH_ENABLED'] = 'true'
        os.environ['PATRIMONIO_USERNAME'] = 'admin'
        os.environ['PATRIMONIO_PASSWORD'] = 'secret'
        os.environ['APP_LOGIN_RATE_LIMIT_ATTEMPTS'] = '2'
        try:
            client = TestClient(app)
            self.assertEqual(client.post('/auth/login', json={'username': 'admin', 'password': 'bad'}).status_code, 401)
            self.assertEqual(client.post('/auth/login', json={'username': 'admin', 'password': 'bad'}).status_code, 401)
            self.assertEqual(client.post('/auth/login', json={'username': 'admin', 'password': 'secret'}).status_code, 429)
        finally:
            auth._LOGIN_ATTEMPTS.clear()
            if previous_enabled is None:
                os.environ.pop('APP_AUTH_ENABLED', None)
            else:
                os.environ['APP_AUTH_ENABLED'] = previous_enabled
            if previous_user is None:
                os.environ.pop('PATRIMONIO_USERNAME', None)
            else:
                os.environ['PATRIMONIO_USERNAME'] = previous_user
            if previous_password is None:
                os.environ.pop('PATRIMONIO_PASSWORD', None)
            else:
                os.environ['PATRIMONIO_PASSWORD'] = previous_password
            if previous_attempts is None:
                os.environ.pop('APP_LOGIN_RATE_LIMIT_ATTEMPTS', None)
            else:
                os.environ['APP_LOGIN_RATE_LIMIT_ATTEMPTS'] = previous_attempts

    def test_export_endpoint_includes_core_tables(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        antonio = crud.create_owner(db, 'Antonio')
        patri = crud.create_owner(db, 'Patri')
        asset = crud.create_asset(db, {'name': 'Export Asset', 'category': 'Cash'})
        crud.upsert_investing_asset(db, 'Cash', False)
        crud.record_price_quote(db, asset, {'provider': 'yahoo', 'symbol': 'SAN.MC', 'price': 4.25, 'currency': 'EUR', 'as_of': '2026-07-18T12:00:00'})
        position = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=1.0, value=100.0, source='manual')
        crud.replace_position_ownership(db, position, [
            {'owner_id': antonio.id, 'share': 0.5},
            {'owner_id': patri.id, 'share': 0.5},
        ])
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.get('/export')
            self.assertEqual(response.status_code, 200)
            self.assertIn('attachment', response.headers.get('content-disposition', ''))
            payload = response.json()
            self.assertEqual(payload['version'], 1)
            self.assertEqual(payload['owners'][0]['name'], 'Antonio')
            self.assertEqual(payload['assets'][0]['name'], 'Export Asset')
            self.assertEqual(payload['investing_assets'][0], {'category': 'Cash', 'is_invested': False})
            self.assertEqual(payload['price_history'][0]['symbol'], 'SAN.MC')
            self.assertEqual(payload['positions'][0]['value'], 100.0)
            self.assertEqual(len(payload['position_ownerships']), 2)
        finally:
            app.dependency_overrides.clear()

    def test_restore_endpoint_requires_confirmation(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/restore', json={'confirm_restore': False, 'backup': {'version': 1}})
            self.assertEqual(response.status_code, 400)
        finally:
            app.dependency_overrides.clear()

    def test_restore_endpoint_replaces_database_from_export(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        antonio = crud.create_owner(db, 'Antonio')
        patri = crud.create_owner(db, 'Patri')
        asset = crud.create_asset(db, {'name': 'Restore Asset', 'category': 'RF'})
        position = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=2.0, value=250.0, source='manual')
        crud.replace_position_ownership(db, position, [
            {'owner_id': antonio.id, 'share': 0.6},
            {'owner_id': patri.id, 'share': 0.4},
        ])
        crud.upsert_investing_asset(db, 'RF', True)
        backup = crud.export_data(db)
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            delete_response = client.delete(f'/assets/{asset.id}')
            self.assertEqual(delete_response.status_code, 204)

            restore_response = client.post('/restore', json={'confirm_restore': True, 'backup': jsonable_backup(backup)})
            self.assertEqual(restore_response.status_code, 200)
            self.assertEqual(restore_response.json()['positions'], 1)

            details = client.get('/dashboard/details?as_of_date=2025-07-01').json()
            self.assertEqual(len(details), 2)
            self.assertEqual(sum(row['value'] for row in details), 250.0)
            self.assertEqual(client.get('/investing-assets/').json(), [{'category': 'RF', 'is_invested': True}])
        finally:
            app.dependency_overrides.clear()

    def test_investing_assets_defaults_are_category_based(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            crud.create_asset(db, {'name': 'Main Home', 'category': 'Casa', 'asset_type': 'real_estate'})
            crud.create_asset(db, {'name': 'Bank Account', 'category': 'CASH', 'asset_type': 'cash'})
            crud.create_asset(db, {'name': 'Bitcoin', 'category': 'Crypto', 'asset_type': 'stock'})

            rows = crud.list_investing_assets(db)
            by_category = {row['category']: row['is_invested'] for row in rows}

            self.assertFalse(by_category['Casa'])
            self.assertFalse(by_category['CASH'])
            self.assertTrue(by_category['Crypto'])
        finally:
            db.close()

    def test_investing_assets_api_persists_category_override(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        crud.create_asset(db, {'name': 'Main Home', 'category': 'Casa', 'asset_type': 'real_estate'})
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.get('/investing-assets/')
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json(), [{'category': 'Casa', 'is_invested': False}])

            update_response = client.put('/investing-assets/Casa', json={'is_invested': True})
            self.assertEqual(update_response.status_code, 200)
            self.assertEqual(update_response.json(), {'category': 'Casa', 'is_invested': True})

            response = client.get('/investing-assets/')
            self.assertEqual(response.json(), [{'category': 'Casa', 'is_invested': True}])
        finally:
            app.dependency_overrides.clear()

    def test_mutation_endpoints_write_audit_log(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/owners/', json={'name': 'Audited Owner', 'type': 'person'})
            self.assertEqual(response.status_code, 200)

            logs = client.get('/audit-log/').json()
            self.assertEqual(logs[0]['action'], 'create_owner')
            self.assertEqual(logs[0]['entity_type'], 'owner')
            self.assertEqual(logs[0]['entity_id'], str(response.json()['id']))
        finally:
            app.dependency_overrides.clear()

    def test_price_history_endpoint_lists_stored_quotes(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        asset = crud.create_asset(db, {'name': 'Quoted Asset', 'category': 'RF'})
        crud.record_price_quote(db, asset, {'provider': 'yahoo', 'symbol': 'SAN.MC', 'price': 4.25, 'currency': 'EUR', 'as_of': '2026-07-18T12:00:00'})
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.get('/prices/history')
            self.assertEqual(response.status_code, 200)
            rows = response.json()
            self.assertEqual(rows[0]['symbol'], 'SAN.MC')
            self.assertEqual(rows[0]['price'], 4.25)
        finally:
            app.dependency_overrides.clear()

    def test_create_position_persists_snapshot(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            asset = models.Asset(name='Test Asset', valuation_method='market_direct')
            db.add(asset)
            db.commit()
            db.refresh(asset)

            position = crud.create_position(
                db,
                asset,
                as_of_date=datetime(2024, 1, 1),
                quantity=3.0,
                value=150.0,
                source='test',
            )

            self.assertEqual(position.asset_id, asset.id)
            self.assertEqual(position.quantity, 3.0)
            self.assertEqual(position.value, 150.0)
            self.assertEqual(db.query(models.Position).count(), 1)
        finally:
            db.close()

    def test_position_api_creates_updates_and_deletes_manual_position(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        owner = crud.create_owner(db, 'Antonio')
        asset = crud.create_asset(db, {'name': 'Manual Asset', 'category': 'RF'})
        owner_id = owner.id
        asset_id = asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            create_response = client.post('/positions/', json={
                'asset_id': asset_id,
                'owner_id': owner_id,
                'as_of_date': '2025-07-01',
                'quantity': 3.0,
                'value': 120.0,
                'broker': 'BKT',
                'source': 'manual',
            })
            self.assertEqual(create_response.status_code, 200)
            created = create_response.json()
            self.assertEqual(created['asset_id'], asset_id)
            self.assertEqual(created['owner_id'], owner_id)
            self.assertEqual(created['broker'], 'BKT')
            self.assertEqual(created['source'], 'manual')

            update_response = client.put(f"/positions/{created['id']}", json={'quantity': 4.0, 'value': 160.0})
            self.assertEqual(update_response.status_code, 200)
            updated = update_response.json()
            self.assertEqual(updated['quantity'], 4.0)
            self.assertEqual(updated['value'], 160.0)

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            self.assertEqual(details_response.status_code, 200)
            details = details_response.json()
            self.assertEqual(details[0]['position_id'], created['id'])
            self.assertEqual(details[0]['asset_id'], asset_id)
            self.assertEqual(details[0]['owner_id'], owner_id)

            delete_response = client.delete(f"/positions/{created['id']}")
            self.assertEqual(delete_response.status_code, 204)
        finally:
            app.dependency_overrides.clear()

    def test_bulk_position_api_saves_snapshot_and_syncs_asset_category(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        owner = crud.create_owner(db, 'Antonio')
        asset = crud.create_asset(db, {'name': 'Manual Asset', 'category': 'RF'})
        owner_id = owner.id
        asset_id = asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            create_response = client.post('/positions/bulk', json={'positions': [{
                'asset_id': asset_id,
                'owner_id': owner_id,
                'as_of_date': '2025-07-01',
                'quantity': 3.0,
                'value': 120.0,
                'category': 'RV',
                'broker': 'BKT',
                'source': 'manual',
            }]})
            self.assertEqual(create_response.status_code, 200)
            created_rows = create_response.json()
            self.assertEqual(len(created_rows), 1)
            position_id = created_rows[0]['id']
            self.assertEqual(created_rows[0]['value'], 120.0)

            assets_response = client.get('/assets/')
            self.assertEqual(assets_response.status_code, 200)
            self.assertEqual(assets_response.json()[0]['category'], 'RV')

            update_response = client.post('/positions/bulk', json={'positions': [{
                'position_id': position_id,
                'asset_id': asset_id,
                'owner_id': owner_id,
                'as_of_date': '2025-07-01',
                'quantity': 4.0,
                'value': 160.0,
                'category': 'RF',
                'broker': 'BKT',
                'source': 'manual',
            }]})
            self.assertEqual(update_response.status_code, 200)
            updated_rows = update_response.json()
            self.assertEqual(updated_rows[0]['id'], position_id)
            self.assertEqual(updated_rows[0]['quantity'], 4.0)
            self.assertEqual(updated_rows[0]['value'], 160.0)

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            self.assertEqual(details_response.status_code, 200)
            details = details_response.json()
            self.assertEqual(len(details), 1)
            self.assertEqual(details[0]['category'], 'RF')

            delete_response = client.post('/positions/bulk', json={'positions': [{
                'position_id': position_id,
                'asset_id': asset_id,
                'owner_id': owner_id,
                'as_of_date': '2025-07-01',
                'quantity': 4.0,
                'value': 0.0,
                'category': 'RF',
                'broker': 'BKT',
                'source': 'manual',
            }]})
            self.assertEqual(delete_response.status_code, 200)
            self.assertEqual(delete_response.json(), [])
            self.assertEqual(client.get('/dashboard/details?as_of_date=2025-07-01').json(), [])
        finally:
            app.dependency_overrides.clear()

    def test_bulk_snapshot_failure_rolls_back_staged_rows(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)
        db = TestingSession()
        existing_asset = crud.create_asset(db, {'name': 'Existing asset', 'category': 'Cash'})
        existing_asset_id = existing_asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/positions/bulk', json={
                'as_of_date': '2025-08-01',
                'replace_snapshot': True,
                'positions': [
                    {
                        'asset_name': 'Should not persist',
                        'category': 'Cash',
                        'as_of_date': '2025-08-01',
                        'quantity': 1,
                        'value': 100,
                    },
                    {
                        'asset_id': existing_asset_id,
                        'owner_id': 99999,
                        'as_of_date': '2025-08-01',
                        'quantity': 1,
                        'value': 100,
                    },
                ],
            })
            self.assertEqual(response.status_code, 404)
            check_db = TestingSession()
            try:
                self.assertIsNone(crud.get_asset_by_name(check_db, 'Should not persist'))
                self.assertEqual(check_db.query(models.Position).count(), 0)
            finally:
                check_db.close()
        finally:
            app.dependency_overrides.clear()

    def test_bulk_position_api_merges_duplicate_snapshot_rows(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        owner = crud.create_owner(db, 'Antonio')
        asset = crud.create_asset(db, {'name': 'BBVA', 'category': 'RF'})
        first = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=1.0, value=100.0, owner=owner, broker='BKT', source='manual')
        second = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=2.0, value=80.0, owner=owner, broker='SAB', source='manual')
        owner_id = owner.id
        asset_id = asset.id
        first_id = first.id
        second_id = second.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/positions/bulk', json={'positions': [
                {
                    'position_id': first_id,
                    'asset_id': asset_id,
                    'owner_id': owner_id,
                    'as_of_date': '2025-07-01',
                    'quantity': 1.0,
                    'value': 100.0,
                    'category': 'RF',
                    'broker': 'BKT',
                    'source': 'manual',
                },
                {
                    'position_id': second_id,
                    'asset_id': asset_id,
                    'owner_id': owner_id,
                    'as_of_date': '2025-07-01',
                    'quantity': 2.0,
                    'value': 80.0,
                    'category': 'RF',
                    'broker': 'BKT',
                    'source': 'manual',
                },
            ]})
            self.assertEqual(response.status_code, 200)
            rows = response.json()
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['id'], first_id)
            self.assertEqual(rows[0]['quantity'], 3.0)
            self.assertEqual(rows[0]['value'], 180.0)
            self.assertEqual(rows[0]['broker'], 'BKT')

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            self.assertEqual(details_response.status_code, 200)
            details = details_response.json()
            self.assertEqual(len(details), 1)
            self.assertEqual(details[0]['position_id'], first_id)
            self.assertEqual(details[0]['quantity'], 3.0)
            self.assertEqual(details[0]['value'], 180.0)
        finally:
            app.dependency_overrides.clear()

    def test_bulk_position_api_defaults_common_position_to_family_split(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        crud.create_owner(db, 'Antonio')
        crud.create_owner(db, 'Patri')
        asset = crud.create_asset(db, {'name': 'Common Cash', 'category': 'Cash'})
        asset_id = asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/positions/bulk', json={'positions': [{
                'asset_id': asset_id,
                'owner_id': None,
                'as_of_date': '2025-07-01',
                'quantity': 1.0,
                'value': 1000.0,
                'category': 'Cash',
                'broker': 'BKT',
                'source': 'manual',
            }]})
            self.assertEqual(response.status_code, 200)

            position_id = response.json()[0]['id']
            ownership_response = client.get(f'/positions/{position_id}/ownership')
            self.assertEqual(ownership_response.status_code, 200)
            owner_map = {row['owner_name']: row['share'] for row in ownership_response.json()}
            self.assertAlmostEqual(owner_map['Antonio'], 0.5)
            self.assertAlmostEqual(owner_map['Patri'], 0.5)

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            self.assertEqual(details_response.status_code, 200)
            details = details_response.json()
            self.assertEqual({row['owner_name'] for row in details}, {'Antonio', 'Patri'})
            self.assertAlmostEqual(sum(row['value'] for row in details), 1000.0)
            self.assertTrue(all(row['value'] == 500.0 for row in details))
        finally:
            app.dependency_overrides.clear()

    def test_bulk_position_api_persists_custom_position_split(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        antonio = crud.create_owner(db, 'Antonio')
        patri = crud.create_owner(db, 'Patri')
        asset = crud.create_asset(db, {'name': 'Common Stock', 'category': 'RF'})
        antonio_id = antonio.id
        patri_id = patri.id
        asset_id = asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/positions/bulk', json={'positions': [{
                'asset_id': asset_id,
                'owner_id': None,
                'as_of_date': '2025-07-01',
                'quantity': 10.0,
                'value': 1000.0,
                'category': 'RF',
                'broker': 'BKT',
                'source': 'manual',
                'ownership_shares': [
                    {'owner_id': antonio_id, 'share': 0.3},
                    {'owner_id': patri_id, 'share': 0.7},
                ],
            }]})
            self.assertEqual(response.status_code, 200)
            position_id = response.json()[0]['id']

            ownership_response = client.get(f'/positions/{position_id}/ownership')
            self.assertEqual(ownership_response.status_code, 200)
            owner_map = {row['owner_name']: row['share'] for row in ownership_response.json()}
            self.assertAlmostEqual(owner_map['Antonio'], 0.3)
            self.assertAlmostEqual(owner_map['Patri'], 0.7)

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            details = details_response.json()
            detail_map = {row['owner_name']: row['value'] for row in details}
            self.assertAlmostEqual(detail_map['Antonio'], 300.0)
            self.assertAlmostEqual(detail_map['Patri'], 700.0)
        finally:
            app.dependency_overrides.clear()

    def test_catalog_api_creates_owner_and_asset_for_manual_entry(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            owner_response = client.post('/owners/', json={'name': 'Empresa Nueva', 'type': 'company'})
            self.assertEqual(owner_response.status_code, 200)
            owner = owner_response.json()
            duplicate_owner_response = client.post('/owners/', json={'name': 'Empresa Nueva', 'type': 'company'})
            self.assertEqual(duplicate_owner_response.status_code, 400)
            update_owner_response = client.put(f"/owners/{owner['id']}", json={'name': 'Empresa Nueva SL', 'type': 'company'})
            self.assertEqual(update_owner_response.status_code, 200)
            self.assertEqual(update_owner_response.json()['name'], 'Empresa Nueva SL')

            asset_response = client.post('/assets/', json={
                'name': 'SAN',
                'category': 'RF',
                'asset_type': 'stock',
                'valuation_method': 'price_provider',
                'price_provider': 'yahoo',
                'price_symbol': 'SAN.MC',
            })
            self.assertEqual(asset_response.status_code, 200)
            asset = asset_response.json()
            self.assertEqual(asset['name'], 'SAN')
            self.assertEqual(asset['category'], 'RF')
            self.assertEqual(asset['asset_type'], 'stock')
            self.assertEqual(asset['valuation_method'], 'price_provider')
            self.assertEqual(asset['price_provider'], 'yahoo')
            self.assertEqual(asset['price_symbol'], 'SAN.MC')
            update_asset_response = client.put(f"/assets/{asset['id']}", json={
                'name': 'SAN updated',
                'category': 'Stocks',
                'asset_type': 'stock',
                'valuation_method': 'price_provider',
                'price_provider': 'yahoo',
                'price_symbol': 'SAN.MC',
            })
            self.assertEqual(update_asset_response.status_code, 200)
            updated_asset = update_asset_response.json()
            self.assertEqual(updated_asset['name'], 'SAN updated')
            self.assertEqual(updated_asset['category'], 'Stocks')
            self.assertEqual(updated_asset['asset_type'], 'stock')
            self.assertEqual(updated_asset['price_symbol'], 'SAN.MC')
        finally:
            app.dependency_overrides.clear()

    def test_asset_api_rejects_duplicate_manual_asset_name(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            payload = {'name': 'Duplicate Asset', 'category': 'RF', 'valuation_method': 'market_direct'}
            self.assertEqual(client.post('/assets/', json=payload).status_code, 200)
            duplicate_response = client.post('/assets/', json=payload)
            self.assertEqual(duplicate_response.status_code, 400)
            self.assertEqual(duplicate_response.json()['detail'], 'Asset already exists')
        finally:
            app.dependency_overrides.clear()

    def test_duplicate_assets_endpoint_reports_existing_duplicate_names(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        db.add(models.Asset(name='Same Asset', category='RF'))
        db.add(models.Asset(name=' same asset ', category='Crypto'))
        db.add(models.Asset(name='Unique Asset', category='Cash'))
        db.commit()
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.get('/assets/duplicates')
            self.assertEqual(response.status_code, 200)
            groups = response.json()
            self.assertEqual(len(groups), 1)
            self.assertEqual(groups[0]['count'], 2)
            self.assertEqual({item['category'] for item in groups[0]['assets']}, {'RF', 'Crypto'})
        finally:
            app.dependency_overrides.clear()

    def test_price_quote_endpoint_validates_asset_price_configuration(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        asset = crud.create_asset(db, {
            'name': 'Manual Asset',
            'category': 'Cash',
            'asset_type': 'cash',
            'valuation_method': 'market_direct',
            'price_provider': 'manual',
        })
        asset_id = asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.get(f'/prices/quote?asset_id={asset_id}')
            self.assertEqual(response.status_code, 400)
            self.assertIn('Manual assets', response.json()['detail'])
        finally:
            app.dependency_overrides.clear()

    def test_bulk_position_save_creates_asset_from_row_name(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        owner = crud.create_owner(db, 'Antonio')
        owner_id = owner.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/positions/bulk', json={
                'positions': [{
                    'as_of_date': '2025-07-01',
                    'asset_name': 'Inline New Asset',
                    'category': 'RF',
                    'owner_id': owner_id,
                    'quantity': 3.0,
                    'value': 900.0,
                    'broker': 'BKT',
                    'source': 'manual',
                }]
            })
            self.assertEqual(response.status_code, 200)
            saved_position = response.json()[0]

            assets_response = client.get('/assets/')
            created_asset = next(asset for asset in assets_response.json() if asset['name'] == 'Inline New Asset')
            self.assertEqual(created_asset['category'], 'RF')
            self.assertEqual(saved_position['asset_id'], created_asset['id'])

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            self.assertEqual(details_response.status_code, 200)
            details = details_response.json()
            self.assertEqual(details[0]['asset_name'], 'Inline New Asset')
            self.assertEqual(details[0]['owner_name'], 'Antonio')
            self.assertAlmostEqual(details[0]['value'], 900.0)
        finally:
            app.dependency_overrides.clear()

    def test_bulk_position_save_replaces_snapshot_and_deletes_omitted_rows(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        owner = crud.create_owner(db, 'Antonio')
        kept_asset = crud.create_asset(db, {'name': 'Kept Asset', 'category': 'RF'})
        omitted_asset = crud.create_asset(db, {'name': 'Omitted Asset', 'category': 'Cash'})
        kept_position = crud.create_position(db, kept_asset, datetime(2025, 7, 1), quantity=1.0, value=100.0, owner=owner, broker='BKT', source='manual')
        crud.create_position(db, omitted_asset, datetime(2025, 7, 1), quantity=1.0, value=200.0, owner=owner, broker='BKT', source='manual')
        crud.create_position(db, omitted_asset, datetime(2025, 8, 1), quantity=1.0, value=300.0, owner=owner, broker='BKT', source='manual')
        owner_id = owner.id
        kept_asset_id = kept_asset.id
        kept_position_id = kept_position.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            response = client.post('/positions/bulk', json={
                'as_of_date': '2025-07-01',
                'replace_snapshot': True,
                'positions': [{
                    'position_id': kept_position_id,
                    'as_of_date': '2025-07-01',
                    'asset_id': kept_asset_id,
                    'category': 'RF',
                    'owner_id': owner_id,
                    'quantity': 2.0,
                    'value': 150.0,
                    'broker': 'BKT',
                    'source': 'manual',
                }]
            })
            self.assertEqual(response.status_code, 200)

            july_details = client.get('/dashboard/details?as_of_date=2025-07-01').json()
            self.assertEqual(len(july_details), 1)
            self.assertEqual(july_details[0]['asset_name'], 'Kept Asset')
            self.assertAlmostEqual(july_details[0]['value'], 150.0)

            august_details = client.get('/dashboard/details?as_of_date=2025-08-01').json()
            self.assertEqual(len(august_details), 1)
            self.assertEqual(august_details[0]['asset_name'], 'Omitted Asset')
        finally:
            app.dependency_overrides.clear()

    def test_catalog_api_deletes_assets_and_blocks_used_owners(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        owner = crud.create_owner(db, 'Temporary Owner')
        unused_owner = crud.create_owner(db, 'Unused Owner')
        asset = crud.create_asset(db, {'name': 'Temporary Asset', 'category': 'Cash'})
        crud.create_position(db, asset, datetime(2025, 7, 1), quantity=1.0, value=100.0, owner=owner, source='manual')
        owner_id = owner.id
        unused_owner_id = unused_owner.id
        asset_id = asset.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            blocked_owner_response = client.delete(f'/owners/{owner_id}')
            self.assertEqual(blocked_owner_response.status_code, 400)

            delete_asset_response = client.delete(f'/assets/{asset_id}')
            self.assertEqual(delete_asset_response.status_code, 204)
            self.assertEqual(client.get('/positions/?as_of_date=2025-07-01').json(), [])

            delete_owner_response = client.delete(f'/owners/{unused_owner_id}')
            self.assertEqual(delete_owner_response.status_code, 204)
        finally:
            app.dependency_overrides.clear()

    def test_position_ownership_api_updates_and_validates_shares(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        antonio = crud.create_owner(db, 'Antonio')
        patri = crud.create_owner(db, 'Patri')
        asset = crud.create_asset(db, {'name': 'Common Account', 'category': 'Cash'})
        position = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=1.0, value=1000.0, source='manual')
        antonio_id = antonio.id
        patri_id = patri.id
        position_id = position.id
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            update_response = client.put(f'/positions/{position_id}/ownership', json={'shares': [
                {'owner_id': antonio_id, 'share': 0.5},
                {'owner_id': patri_id, 'share': 0.5},
            ]})
            self.assertEqual(update_response.status_code, 200)
            rows = update_response.json()
            self.assertEqual(len(rows), 2)
            self.assertAlmostEqual(sum(row['share'] for row in rows), 1.0)

            get_response = client.get(f'/positions/{position_id}/ownership')
            self.assertEqual(get_response.status_code, 200)
            owner_map = {row['owner_name']: row['share'] for row in get_response.json()}
            self.assertAlmostEqual(owner_map['Antonio'], 0.5)
            self.assertAlmostEqual(owner_map['Patri'], 0.5)

            too_much_response = client.put(f'/positions/{position_id}/ownership', json={'shares': [
                {'owner_id': antonio_id, 'share': 0.7},
                {'owner_id': patri_id, 'share': 0.5},
            ]})
            self.assertEqual(too_much_response.status_code, 400)

            partial_response = client.put(f'/positions/{position_id}/ownership', json={'shares': [
                {'owner_id': antonio_id, 'share': 0.4},
                {'owner_id': patri_id, 'share': 0.4},
            ]})
            self.assertEqual(partial_response.status_code, 200)
            self.assertAlmostEqual(sum(row['share'] for row in partial_response.json()), 0.8)

            all_response = client.get('/positions/ownership')
            self.assertEqual(all_response.status_code, 200)
            all_rows = all_response.json()
            self.assertEqual(all_rows[0]['position_id'], position_id)
            self.assertEqual(len(all_rows[0]['shares']), 2)
        finally:
            app.dependency_overrides.clear()

    def test_same_asset_positions_can_have_different_ownership_splits(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            antonio = crud.create_owner(db, 'Antonio')
            patri = crud.create_owner(db, 'Patri')
            asset = crud.create_asset(db, {'name': 'BBVA', 'category': 'RF'})

            first = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=10.0, value=1000.0, broker='BKT', source='manual')
            second = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=20.0, value=2000.0, broker='SAB', source='manual')
            crud.replace_position_ownership(db, first, [
                {'owner_id': antonio.id, 'share': 0.5},
                {'owner_id': patri.id, 'share': 0.5},
            ])
            crud.replace_position_ownership(db, second, [
                {'owner_id': antonio.id, 'share': 0.25},
                {'owner_id': patri.id, 'share': 0.75},
            ])

            summary = crud.get_dashboard_summary(db, as_of_date=datetime(2025, 7, 1))
            owner_map = {row['owner_name']: row['value'] for row in summary['by_owner']}

            self.assertAlmostEqual(owner_map['Antonio'], 1000.0)
            self.assertAlmostEqual(owner_map['Patri'], 2000.0)
            self.assertAlmostEqual(summary['total_value'], 3000.0)
        finally:
            db.close()

    def test_legacy_asset_ownership_migrates_to_ownerless_positions(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            antonio = crud.create_owner(db, 'Antonio')
            patri = crud.create_owner(db, 'Patri')
            asset = crud.create_asset(db, {'name': 'Legacy Common', 'category': 'Cash'})
            crud.set_ownership(db, asset, antonio, share=0.6)
            crud.set_ownership(db, asset, patri, share=0.4)
            position = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=1.0, value=1000.0, source='legacy')

            migrated = crud.migrate_legacy_asset_ownership_to_positions(db)

            self.assertEqual(migrated, 1)
            shares = {row['owner_name']: row['share'] for row in crud.get_position_ownership(db, position)}
            self.assertAlmostEqual(shares['Antonio'], 0.6)
            self.assertAlmostEqual(shares['Patri'], 0.4)
            self.assertEqual(db.query(models.Ownership).count(), 0)
        finally:
            db.close()

    def test_create_position_separates_owner_and_broker_holdings(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            antonio = crud.create_owner(db, 'Antonio')
            patri = crud.create_owner(db, 'Patri')
            asset = models.Asset(name='BBVA', category='RF', valuation_method='market_direct')
            db.add(asset)
            db.commit()
            db.refresh(asset)

            crud.create_position(db, asset, datetime(2024, 1, 1), quantity=10.0, value=100.0, owner=antonio, broker='BKT', source='test')
            crud.create_position(db, asset, datetime(2024, 1, 1), quantity=20.0, value=200.0, owner=patri, broker='SAB', source='test')

            positions = crud.get_positions(db, as_of_date=datetime(2024, 1, 1))

            self.assertEqual(len(positions), 2)
            summary = crud.get_dashboard_summary(db, as_of_date=datetime(2024, 1, 1))
            owner_map = {row['owner_name']: row['value'] for row in summary['by_owner']}
            self.assertAlmostEqual(owner_map['Antonio'], 100.0)
            self.assertAlmostEqual(owner_map['Patri'], 200.0)
            broker_map = {row['broker']: row['value'] for row in summary['by_broker']}
            self.assertAlmostEqual(broker_map['BKT'], 100.0)
            self.assertAlmostEqual(broker_map['SAB'], 200.0)
        finally:
            db.close()

    def test_dashboard_summary_aggregates_owner_and_asset_totals(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            owner_a = crud.create_owner(db, 'Antonio')
            owner_b = crud.create_owner(db, 'Patri')

            cash = models.Asset(name='Cash', category='cash', valuation_method='market_direct')
            house = models.Asset(name='House', category='real_estate', valuation_method='market_minus_debt')
            db.add_all([cash, house])
            db.commit()
            db.refresh(cash)
            db.refresh(house)

            cash_position = crud.create_position(db, cash, datetime(2024, 1, 1), quantity=1.0, value=100.0, source='test')
            crud.replace_position_ownership(db, cash_position, [
                {'owner_id': owner_a.id, 'share': 0.5},
                {'owner_id': owner_b.id, 'share': 0.5},
            ])
            crud.create_position(db, house, datetime(2024, 1, 1), quantity=1.0, value=200.0, owner=owner_a, source='test')

            summary = crud.get_dashboard_summary(db, as_of_date=datetime(2024, 1, 1))

            self.assertAlmostEqual(summary['total_value'], 300.0)
            owner_map = {row['owner_name']: row['value'] for row in summary['by_owner']}
            self.assertAlmostEqual(owner_map['Antonio'], 250.0)
            self.assertAlmostEqual(owner_map['Patri'], 50.0)
            asset_map = {row['asset_name']: row['value'] for row in summary['by_asset']}
            self.assertAlmostEqual(asset_map['Cash'], 100.0)
            self.assertAlmostEqual(asset_map['House'], 200.0)
            category_map = {row['category']: row['value'] for row in summary['by_category']}
            self.assertAlmostEqual(category_map['cash'], 100.0)
            self.assertAlmostEqual(category_map['real_estate'], 200.0)
        finally:
            db.close()

    def test_dashboard_uses_family_attributable_value_for_partial_shared_asset(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            antonio = crud.create_owner(db, 'Antonio')
            patri = crud.create_owner(db, 'Patri')
            third_party = crud.create_owner(db, 'Third Party')
            company = models.Asset(name='Company Asset', category='Empresa', valuation_method='company_net_assets')
            db.add(company)
            db.commit()
            db.refresh(company)

            position = crud.create_position(db, company, datetime(2024, 1, 1), quantity=1.0, value=100000.0, source='test')
            crud.replace_position_ownership(db, position, [
                {'owner_id': antonio.id, 'share': 0.4},
                {'owner_id': patri.id, 'share': 0.4},
                {'owner_id': third_party.id, 'share': 0.2},
            ])

            summary = crud.get_dashboard_summary(db, as_of_date=datetime(2024, 1, 1))
            details = crud.get_dashboard_details(db, as_of_date=datetime(2024, 1, 1))

            self.assertAlmostEqual(summary['total_value'], 80000.0)
            owner_map = {row['owner_name']: row['value'] for row in summary['by_owner']}
            self.assertAlmostEqual(owner_map['Antonio'], 40000.0)
            self.assertAlmostEqual(owner_map['Patri'], 40000.0)
            self.assertAlmostEqual(owner_map['Third Party'], 0.0)
            asset_map = {row['asset_name']: row['value'] for row in summary['by_asset']}
            self.assertAlmostEqual(asset_map['Company Asset'], 80000.0)
            self.assertEqual({row['owner_name'] for row in details}, {'Antonio', 'Patri'})
            self.assertAlmostEqual(sum(row['value'] for row in details), 80000.0)
        finally:
            db.close()

    def test_dashboard_looks_through_company_ownership_chain(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            antonio = crud.create_owner(db, 'Antonio', is_family_member=True)
            patri = crud.create_owner(db, 'Patri', is_family_member=True)
            hereditas = crud.create_owner(db, 'Hereditas', type='company')
            subsidiary = crud.create_owner(db, 'Other Company', type='company')
            asset = crud.create_asset(db, {'name': 'Subsidiary cash', 'category': 'Cash'})
            as_of_date = datetime(2025, 1, 1)
            crud.create_position(
                db,
                asset,
                as_of_date,
                quantity=1.0,
                value=1000000.0,
                owner=subsidiary,
                source='test',
            )
            crud.create_entity_ownership(db, {
                'owner_id': antonio.id,
                'owned_id': hereditas.id,
                'share': 0.5,
                'effective_from': as_of_date,
            })
            crud.create_entity_ownership(db, {
                'owner_id': patri.id,
                'owned_id': hereditas.id,
                'share': 0.5,
                'effective_from': as_of_date,
            })
            crud.create_entity_ownership(db, {
                'owner_id': hereditas.id,
                'owned_id': subsidiary.id,
                'share': 0.75,
                'effective_from': as_of_date,
            })

            summary = crud.get_dashboard_summary(db, as_of_date=as_of_date)
            details = crud.get_dashboard_details(db, as_of_date=as_of_date)
            owner_map = {row['owner_name']: row['value'] for row in summary['by_owner']}

            self.assertAlmostEqual(summary['total_value'], 750000.0)
            self.assertAlmostEqual(owner_map['Antonio'], 375000.0)
            self.assertAlmostEqual(owner_map['Patri'], 375000.0)
            self.assertEqual({row['owner_name'] for row in details}, {'Antonio', 'Patri'})
            self.assertAlmostEqual(sum(row['value'] for row in details), 750000.0)
        finally:
            db.close()

    def test_position_snapshot_api_returns_raw_ownerless_position_for_editing(self):
        engine = create_engine(
            'sqlite://',
            connect_args={'check_same_thread': False},
            poolclass=StaticPool,
        )
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        antonio = crud.create_owner(db, 'Antonio')
        patri = crud.create_owner(db, 'Patri')
        asset = crud.create_asset(db, {'name': 'Common Account', 'category': 'Cash'})
        position = crud.create_position(db, asset, datetime(2025, 7, 1), quantity=1.0, value=1000.0, owner=None, broker='BKT', source='manual')
        crud.replace_position_ownership(db, position, [
            {'owner_id': antonio.id, 'share': 0.5},
            {'owner_id': patri.id, 'share': 0.5},
        ])
        db.close()

        def override_get_db():
            session = TestingSession()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        client = TestClient(app)
        try:
            raw_response = client.get('/positions/snapshot?as_of_date=2025-07-01')
            self.assertEqual(raw_response.status_code, 200)
            raw_rows = raw_response.json()
            self.assertEqual(len(raw_rows), 1)
            self.assertIsNone(raw_rows[0]['owner_id'])
            self.assertIsNone(raw_rows[0]['owner_name'])
            self.assertEqual(raw_rows[0]['value'], 1000.0)

            details_response = client.get('/dashboard/details?as_of_date=2025-07-01')
            self.assertEqual(details_response.status_code, 200)
            details = details_response.json()
            self.assertEqual(len(details), 2)
            self.assertEqual({row['owner_name'] for row in details}, {'Antonio', 'Patri'})
            self.assertAlmostEqual(sum(row['value'] for row in details), 1000.0)
        finally:
            app.dependency_overrides.clear()

    def test_get_positions_uses_exact_selected_date(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            asset = models.Asset(name='Test Asset', valuation_method='market_direct')
            db.add(asset)
            db.commit()
            db.refresh(asset)

            crud.create_position(db, asset, datetime(2024, 1, 1), quantity=1.0, value=100.0, source='test')
            crud.create_position(db, asset, datetime(2024, 2, 1), quantity=2.0, value=200.0, source='test')

            positions = crud.get_positions(db, as_of_date=datetime(2024, 1, 15))

            self.assertEqual(len(positions), 0)
        finally:
            db.close()

    def test_get_available_dates_returns_distinct_snapshot_dates(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            asset_a = models.Asset(name='Asset A', valuation_method='market_direct')
            asset_b = models.Asset(name='Asset B', valuation_method='market_direct')
            db.add_all([asset_a, asset_b])
            db.commit()
            db.refresh(asset_a)
            db.refresh(asset_b)

            crud.create_position(db, asset_a, datetime(2024, 1, 1), quantity=1.0, value=100.0, source='test')
            crud.create_position(db, asset_b, datetime(2024, 2, 1), quantity=2.0, value=200.0, source='test')
            crud.create_position(db, asset_a, datetime(2024, 2, 1), quantity=3.0, value=300.0, source='test')

            available_dates = crud.get_available_dates(db)

            self.assertEqual(available_dates, [datetime(2024, 1, 1), datetime(2024, 2, 1)])
        finally:
            db.close()

    def test_import_mapping_uses_excel_headers(self):
        row = {'Tipo Activo': 'RF', 'Activo': 'BBVA', 'Cantidad': 792.0, 'NAV': 1234.5, 'Broker': 'BBVA'}

        self.assertEqual(detect_name(row), 'BBVA')
        self.assertEqual(detect_category(row), 'RF')
        self.assertEqual(detect_net_asset_value(row), 1234.5)

    def test_sheet_name_date_wins_over_stale_header_date(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '2025-07-05'
        sheet['D2'] = 'Date'
        sheet['E2'] = datetime(2025, 1, 6)

        detected = detect_sheet_date(sheet, sheet.title, ['Tipo Activo', 'Activo', 'Cantidad', 'NAV'])

        self.assertEqual(detected, datetime(2025, 7, 5))

    def test_import_file_processes_detail_sheets_with_snapshot_date(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                path = os.path.join(temp_dir, 'Patrimonio Patri.xlsx')
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = '2025-06-01'
                sheet['A1'] = 'Patrimonio Patri'
                sheet['D2'] = 'Date'
                sheet['E2'] = datetime(2025, 6, 1)
                sheet['A4'] = 'Tipo Activo'
                sheet['B4'] = 'Activo'
                sheet['C4'] = 'Cantidad'
                sheet['D4'] = 'NAV'
                sheet['E4'] = 'Broker'
                sheet['A5'] = 'RF'
                sheet['B5'] = 'BBVA'
                sheet['C5'] = 792.0
                sheet['D5'] = 1234.5
                sheet['E5'] = 'BBVA'
                sheet['A6'] = 'RF'
                sheet['B6'] = 'IAG'
                sheet['C6'] = 650.0
                sheet['D6'] = 1294.48
                sheet['E6'] = 'BBVA'
                workbook.save(path)

                import_file(path, db, dry_run=False)

            assets = db.query(models.Asset).order_by(models.Asset.name).all()
            self.assertEqual(len(assets), 2)
            asset_map = {asset.name: asset for asset in assets}

            bbva = asset_map['BBVA']
            position = db.query(models.Position).filter(models.Position.asset_id == bbva.id).first()
            self.assertEqual(position.as_of_date, datetime(2025, 6, 1))
            self.assertEqual(position.owner.name, 'Patri')
            self.assertEqual(position.broker, 'BBVA')
            self.assertAlmostEqual(position.value, 1234.5)
        finally:
            db.close()

    def test_dashboard_history_returns_summary_and_details_for_each_date(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            owner = crud.create_owner(db, 'Antonio')
            asset = models.Asset(name='Asset A', category='RF', valuation_method='market_direct')
            db.add(asset)
            db.commit()
            db.refresh(asset)

            crud.create_position(db, asset, datetime(2024, 1, 1), quantity=1.0, value=100.0, owner=owner, broker='BKT', source='test')
            crud.create_position(db, asset, datetime(2024, 2, 1), quantity=2.0, value=200.0, owner=owner, broker='BKT', source='test')

            history = crud.get_dashboard_history(db)

            self.assertEqual([point['date'] for point in history], ['2024-01-01', '2024-02-01'])
            self.assertEqual(history[0]['summary']['total_value'], 100.0)
            self.assertEqual(history[1]['summary']['total_value'], 200.0)
            self.assertEqual(history[0]['details'][0]['asset_id'], asset.id)
            self.assertEqual(history[0]['details'][0]['owner_id'], owner.id)
            self.assertIn('position_id', history[0]['details'][0])
        finally:
            db.close()

    def test_dashboard_details_defaults_to_latest_snapshot_only(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            owner = crud.create_owner(db, 'Antonio')
            asset = models.Asset(name='Test Asset', valuation_method='market_direct')
            db.add(asset)
            db.commit()
            db.refresh(asset)

            crud.create_position(db, asset, datetime(2024, 1, 1), quantity=1.0, value=100.0, owner=owner, source='test')
            crud.create_position(db, asset, datetime(2024, 2, 1), quantity=2.0, value=200.0, owner=owner, source='test')

            rows = crud.get_dashboard_details(db)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['quantity'], 2.0)
            self.assertEqual(rows[0]['value'], 200.0)
        finally:
            db.close()

    def test_dashboard_summary_includes_known_owners_without_positions_on_date(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            antonio = crud.create_owner(db, 'Antonio')
            crud.create_owner(db, 'Patri')
            asset = models.Asset(name='Test Asset', valuation_method='market_direct')
            db.add(asset)
            db.commit()
            db.refresh(asset)

            crud.create_position(db, asset, datetime(2024, 1, 1), quantity=1.0, value=100.0, owner=antonio, source='test')

            summary = crud.get_dashboard_summary(db, as_of_date=datetime(2024, 1, 1))
            owner_map = {row['owner_name']: row['value'] for row in summary['by_owner']}

            self.assertEqual(set(owner_map.keys()), {'Antonio', 'Patri'})
            self.assertAlmostEqual(owner_map['Antonio'], 100.0)
            self.assertAlmostEqual(owner_map['Patri'], 0.0)
        finally:
            db.close()

    def test_import_file_skips_zero_nav_rows(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                path = os.path.join(temp_dir, 'Patrimonio ARS.xlsx')
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = '2025-06-01'
                sheet['A1'] = 'Tipo Activo'
                sheet['B1'] = 'Activo'
                sheet['C1'] = 'Cantidad'
                sheet['D1'] = 'NAV'
                sheet['E1'] = 'Broker'
                sheet['A2'] = 'RF'
                sheet['B2'] = 'BBVA'
                sheet['C2'] = 10.0
                sheet['D2'] = 0.0
                sheet['E2'] = 'BKT'
                workbook.save(path)

                import_file(path, db, dry_run=False)

            self.assertEqual(db.query(models.Position).count(), 0)
        finally:
            db.close()

    def test_import_file_does_not_create_position_when_holding_disappears(self):
        engine = create_engine('sqlite:///:memory:')
        TestingSession = sessionmaker(bind=engine)
        models.Base.metadata.create_all(bind=engine)

        db = TestingSession()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                path = os.path.join(temp_dir, 'Patrimonio ARS.xlsx')
                workbook = Workbook()
                later_sheet = workbook.active
                later_sheet.title = '2025-07-01'
                later_sheet['A1'] = 'Tipo Activo'
                later_sheet['B1'] = 'Activo'
                later_sheet['C1'] = 'Cantidad'
                later_sheet['D1'] = 'NAV'
                later_sheet['E1'] = 'Broker'
                later_sheet['A2'] = 'RF'
                later_sheet['B2'] = 'IAG'
                later_sheet['C2'] = 5.0
                later_sheet['D2'] = 50.0
                later_sheet['E2'] = 'BKT'

                earlier_sheet = workbook.create_sheet('2025-06-01')
                earlier_sheet['A1'] = 'Tipo Activo'
                earlier_sheet['B1'] = 'Activo'
                earlier_sheet['C1'] = 'Cantidad'
                earlier_sheet['D1'] = 'NAV'
                earlier_sheet['E1'] = 'Broker'
                earlier_sheet['A2'] = 'RF'
                earlier_sheet['B2'] = 'BBVA'
                earlier_sheet['C2'] = 10.0
                earlier_sheet['D2'] = 100.0
                earlier_sheet['E2'] = 'BKT'
                workbook.save(path)

                import_file(path, db, dry_run=False)

            bbva = db.query(models.Asset).filter(models.Asset.name == 'BBVA').first()
            june_position = db.query(models.Position).filter(
                models.Position.asset_id == bbva.id,
                models.Position.as_of_date == datetime(2025, 6, 1),
            ).first()
            july_position = db.query(models.Position).filter(
                models.Position.asset_id == bbva.id,
                models.Position.as_of_date == datetime(2025, 7, 1),
            ).first()

            self.assertIsNotNone(june_position)
            self.assertIsNone(july_position)
        finally:
            db.close()


if __name__ == '__main__':
    unittest.main()
