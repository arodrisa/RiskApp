"""
Import Excel files into the asset dimension + valuation history model.
"""
import argparse
import os
from datetime import datetime
from openpyxl import load_workbook

from app.database import SessionLocal, engine
from app.models import Base
from app import crud

COMMON_NAME_KEYS = ['Activo', 'Name', 'Concepto', 'Concept', 'Instrumento', 'Cuenta', 'concepto', '']
COMMON_CATEGORY_KEYS = ['Tipo Activo', 'Tipo de Activo', 'Tipo', 'Category', 'Categoria', 'Asset Type']
COMMON_QUANTITY_KEYS = ['Cantidad', 'quantity', 'qty', 'Quantity']
COMMON_BROKER_KEYS = ['Broker', 'broker', 'Bank', 'bank', 'Banco', 'Broker/Bank']
COMMON_NAV_KEYS = ['NAV', 'Net Asset Value', 'net_asset_value', 'Net asset value', 'Valor NAV']


def new_import_report(file_name):
    return {
        'file': file_name,
        'skipped_file': False,
        'sheets_seen': 0,
        'sheets_imported': 0,
        'rows_seen': 0,
        'rows_imported': 0,
        'rows_skipped_missing_category': 0,
        'rows_skipped_empty': 0,
        'holdings': {},
    }


def track_holding(report, asset_name, owner_name, broker):
    key = str(asset_name).strip().lower()
    context = (owner_name or 'Unknown', broker or '(blank)')
    report['holdings'].setdefault(key, {'name': asset_name, 'contexts': set()})
    report['holdings'][key]['contexts'].add(context)


def duplicate_contexts_from_report(report):
    duplicates = []
    for holding in report['holdings'].values():
        contexts = sorted(holding['contexts'])
        if len(contexts) > 1:
            duplicates.append({'asset_name': holding['name'], 'contexts': contexts})
    return sorted(duplicates, key=lambda item: str(item['asset_name']).lower())


def print_import_report(report):
    if report['skipped_file']:
        print(f" Summary for {report['file']}: skipped shared workbook")
        return

    print(
        f" Summary for {report['file']}: "
        f"sheets imported={report['sheets_imported']}, "
        f"rows imported={report['rows_imported']}, "
        f"empty rows skipped={report['rows_skipped_empty']}, "
        f"missing-category rows skipped={report['rows_skipped_missing_category']}"
    )
    duplicates = duplicate_contexts_from_report(report)
    if duplicates:
        print(' Duplicate asset-name contexts:')
        for duplicate in duplicates:
            contexts = ', '.join(f'{owner}/{broker}' for owner, broker in duplicate['contexts'])
            print(f"  - {duplicate['asset_name']}: {contexts}")


def reset_database_for_dev(enabled=False):
    if not enabled:
        Base.metadata.create_all(bind=engine)
        db = SessionLocal()
        try:
            crud.migrate_legacy_asset_ownership_to_positions(db)
        finally:
            db.close()
        return

    environment = os.environ.get('PATRIMONIO_ENV', 'development').lower()
    if environment not in ('dev', 'development', 'local', 'test'):
        raise RuntimeError('Refusing to reset database outside a development environment')

    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def find_table(ws, start_row=1):
    max_row = ws.max_row
    header_row = None

    for r in range(start_row, max_row + 1):
        row = [c.value for c in ws[r]]
        if not any(cell is not None for cell in row):
            continue

        normalized_values = [str(cell).strip() if cell is not None else '' for cell in row]
        has_named_headers = any(
            value and normalize_key(value) in {
                normalize_key(key) for key in COMMON_NAME_KEYS + COMMON_CATEGORY_KEYS + COMMON_QUANTITY_KEYS + COMMON_BROKER_KEYS + COMMON_NAV_KEYS
            }
            for value in normalized_values
        )
        if has_named_headers:
            header_row = r
            break

    if header_row is None:
        for r in range(start_row, max_row + 1):
            row = [c.value for c in ws[r]]
            if not any(cell is not None for cell in row):
                continue
            if any(try_parse_date(cell) is not None for cell in row):
                header_row = r
                break

    if header_row is None:
        return None

    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[header_row]]
    data = []
    r = header_row + 1
    while r <= max_row:
        rowvals = [c.value for c in ws[r]]
        if all(cell is None for cell in rowvals):
            break
        rowdict = {headers[i]: rowvals[i] for i in range(len(headers))}
        data.append(rowdict)
        r += 1
    return headers, data


def inspect_file(path):
    wb = load_workbook(path, data_only=True)
    print(f'File: {path}')
    for name in wb.sheetnames:
        ws = wb[name]
        print(' Sheet:', name, 'rows=', ws.max_row, 'cols=', ws.max_column)
        tbl = find_table(ws, start_row=1)
        if tbl:
            headers, data = tbl
            print('  headers:', headers)
            for r in data[:5]:
                print('   ', r)


def normalize_key(value):
    return str(value).strip().lower().replace(' ', '').replace('_', '')


def get_row_value(row, keys):
    for key in keys:
        if key in row and row[key] not in (None, ''):
            return row[key]
    normalized_lookup = {normalize_key(k): k for k in row.keys() if k is not None}
    for key in keys:
        match_key = normalized_lookup.get(normalize_key(key))
        if match_key is not None and row[match_key] not in (None, ''):
            return row[match_key]
    return None


def detect_name(row):
    value = get_row_value(row, COMMON_NAME_KEYS)
    if value is not None:
        return str(value)
    for v in row.values():
        if isinstance(v, str) and v.strip():
            return str(v)
    for v in row.values():
        if v is not None:
            return str(v)
    return 'Unknown'


def try_parse_date(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        text = str(value).strip()
        if text.endswith(' 00:00:00'):
            text = text[:-9]
        return datetime.strptime(text, '%Y-%m-%d')
    except Exception:
        return None


def detect_quantity(row):
    value = get_row_value(row, COMMON_QUANTITY_KEYS)
    if value not in (None, ''):
        try:
            return float(value)
        except Exception:
            return 1.0
    return 1.0


def detect_broker(row):
    value = get_row_value(row, COMMON_BROKER_KEYS)
    if value not in (None, ''):
        return str(value)
    return None


def detect_category(row):
    value = get_row_value(row, COMMON_CATEGORY_KEYS)
    if value not in (None, ''):
        return str(value)
    return None


def detect_net_asset_value(row):
    value = get_row_value(row, COMMON_NAV_KEYS)
    if value not in (None, ''):
        try:
            return float(value)
        except Exception:
            return 0.0
    return 0.0


def detect_sheet_date(ws, sheet_name, headers):
    parsed_sheet_date = try_parse_date(sheet_name)
    if parsed_sheet_date is not None:
        return parsed_sheet_date

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True):
        for cell in row:
            parsed_date = try_parse_date(cell)
            if parsed_date is not None:
                return parsed_date

    for header in headers:
        parsed_date = try_parse_date(header)
        if parsed_date is not None:
            return parsed_date
    return None


def importable_sheet_contexts(wb):
    contexts = []
    fallback_order = 0

    for sheet_name in wb.sheetnames:
        fallback_order += 1
        if str(sheet_name).strip().lower() == 'resumen':
            continue

        ws = wb[sheet_name]
        tbl = find_table(ws, 1)
        if not tbl:
            continue
        headers, rows = tbl
        if not rows:
            continue

        snapshot_date = detect_sheet_date(ws, sheet_name, headers)
        contexts.append({
            'sheet_name': sheet_name,
            'worksheet': ws,
            'headers': headers,
            'rows': rows,
            'snapshot_date': snapshot_date,
            'fallback_order': fallback_order,
        })

    return sorted(
        contexts,
        key=lambda item: (
            item['snapshot_date'] is None,
            item['snapshot_date'] or datetime.max,
            item['fallback_order'],
        ),
    )


def import_file(path, db, dry_run=True):
    wb = load_workbook(path, data_only=True)
    fname = os.path.basename(path)
    report = new_import_report(fname)
    is_comun = 'Comun' in fname or 'COMUN' in fname
    owner_name = None
    if not is_comun:
        if 'ARS' in fname:
            owner_name = 'Antonio'
        elif 'Patri' in fname:
            owner_name = 'Patri'
        else:
            owner_name = os.path.splitext(fname)[0]

    if is_comun:
        print(f' Skipping shared workbook {fname} as requested')
        report['skipped_file'] = True
        print_import_report(report)
        return report

    if owner_name:
        owner = crud.get_owner_by_name(db, owner_name)
        if not owner:
            owner = crud.create_owner(db, name=owner_name)

    imported_any = False
    for sheet_context in importable_sheet_contexts(wb):
        sheet_name = sheet_context['sheet_name']
        headers = sheet_context['headers']
        rows = sheet_context['rows']
        snapshot_date = sheet_context['snapshot_date']
        report['sheets_seen'] += 1

        print(f' Importing from {fname} sheet {sheet_name}: {len(rows)} rows')
        imported_any = True
        report['sheets_imported'] += 1

        if dry_run:
            for r in rows[:10]:
                print('  ', {k: r[k] for k in list(r.keys())[:6]})
            continue

        has_date_headers = any(try_parse_date(header) is not None for header in headers)

        for r in rows:
            report['rows_seen'] += 1
            if all(value in (None, '') for value in r.values()):
                report['rows_skipped_empty'] += 1
                continue
            if detect_category(r) in (None, ''):
                report['rows_skipped_missing_category'] += 1
                continue

            name = detect_name(r)
            broker = detect_broker(r)
            track_holding(report, name, owner_name, broker)
            asset_data = {
                'name': name,
                'category': detect_category(r),
                'asset_type': r.get('Tipo') or r.get('Type') or None,
                'valuation_method': 'market_minus_debt' if any(x in name.lower() for x in ['casa', 'inmueble']) else 'company_net_assets' if 'heredit' in name.lower() else 'market_direct',
                'is_shared': is_comun,
            }
            asset = crud.create_asset(db, asset_data)

            if has_date_headers:
                for header in headers:
                    parsed_date = try_parse_date(header)
                    if not parsed_date:
                        continue
                    val = r.get(header)
                    if val in (None, ''):
                        continue
                    try:
                        numeric_value = float(val)
                    except Exception:
                        continue
                    if numeric_value <= 0:
                        continue
                    crud.create_position(
                        db,
                        asset,
                        as_of_date=parsed_date,
                        quantity=detect_quantity(r),
                        value=numeric_value,
                        owner=owner,
                        broker=broker,
                        source='import',
                    )
                    report['rows_imported'] += 1
            elif snapshot_date is not None:
                value = detect_net_asset_value(r)
                if value not in (None, '') and value > 0:
                    crud.create_position(
                        db,
                        asset,
                        as_of_date=snapshot_date,
                        quantity=detect_quantity(r),
                        value=value,
                        owner=owner,
                        broker=broker,
                        source='import',
                    )
                    report['rows_imported'] += 1

            db.commit()

    if not imported_any:
        print('  No importable tables found')
    print_import_report(report)
    return report


def main(path, dry_run=True, do_import=False, reset_db_dev=False):
    files = [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.xlsx')]
    if not files:
        print('No .xlsx files found in', path)
        return
    reset_database_for_dev(enabled=do_import and reset_db_dev)
    db = SessionLocal()
    for f in files:
        print('Processing', f)
        if do_import:
            import_file(f, db, dry_run=not do_import)
        else:
            inspect_file(f)
    db.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--path', default='.', help='Path with Excel files')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--do-import', action='store_true')
    parser.add_argument('--keep-existing-db', action='store_true', help='Do not rebuild the database before importing')
    args = parser.parse_args()
    main(args.path, dry_run=args.dry_run, do_import=args.do_import, reset_db_dev=not args.keep_existing_db)
